#!/usr/bin/env python3
"""
hotlist_to_excel.py — 将热榜数据保存为Excel文件
用法：python3.10 hotlist_to_excel.py [--input /tmp/hotlist_data.json] [--output /path/to/file.xlsx]
"""

import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("错误: openpyxl 未安装，请先 pip install openpyxl", file=sys.stderr)
    sys.exit(1)

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i", default="/tmp/hotlist_data.json", help="热榜JSON数据文件")
    parser.add_argument("--output", "-o", default="", help="输出Excel路径（默认保存到Windows桌面）")
    args = parser.parse_args()

    # 读取数据
    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        today = datetime.now().strftime("%Y-%m-%d")
        output_path = f"/mnt/c/Users/yingm/OneDrive/Desktop/热榜数据_{today}.xlsx"

    # 创建Excel
    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    def normalize_engagement(platform, item):
        """归一化各平台互动量到0-100分
        
        使不同平台的互动量可比（微博热度=几百万，B站播放=几万，统一到0-100）
        """
        import math
        raw = 0
        if platform == "zhihu":
            raw = max(
                int(item.get("heat", 0) or 0),
                int(item.get("answer_count", 0) or 0) * 10,
            )
        elif platform == "weibo":
            raw = int(item.get("heat", 0) or 0)
        elif platform == "bilibili":
            raw = max(
                int(item.get("view", 0) or 0),
                int(item.get("like", 0) or 0) * 10,
                int(item.get("reply", 0) or 0) * 50,
            )
        elif platform == "36kr":
            raw = max(
                int(item.get("read", 0) or 0),
                int(item.get("like", 0) or 0) * 50,
            )
        elif platform == "baidu":
            raw = int(item.get("heat", 0) or 0)
        else:
            raw = 0
        if raw <= 0:
            return 0, raw
        ref_max = 10000000  # 1千万参考值
        score = min(100, round(math.log(1 + raw) / math.log(1 + ref_max) * 100, 1))
        return score, raw

    platform_config = {
        "zhihu": {
            "name": "知乎热榜",
            "columns": ["排名", "标题", "热度", "回答数", "互动分", "engagement_score", "摘要", "链接"],
            "extract": lambda item, pf="zhihu": [
                item.get("rank"), item.get("title"), item.get("heat"),
                item.get("answer_count"),
                *normalize_engagement(pf, item),
                item.get("excerpt"), item.get("url")
            ]
        },
        "weibo": {
            "name": "微博热搜",
            "columns": ["排名", "标题", "热度", "标签", "互动分", "engagement_score", "链接"],
            "extract": lambda item, pf="weibo": [
                item.get("rank"), item.get("title"), item.get("heat"),
                item.get("label"),
                *normalize_engagement(pf, item),
                item.get("url")
            ]
        },
        "bilibili": {
            "name": "B站热门",
            "columns": ["排名", "标题", "UP主", "分区", "播放量", "点赞", "评论", "互动分", "engagement_score", "链接"],
            "extract": lambda item, pf="bilibili": [
                item.get("rank"), item.get("title"), item.get("author"),
                item.get("tname"), item.get("view"), item.get("like"),
                item.get("reply"),
                *normalize_engagement(pf, item),
                item.get("url")
            ]
        },
        "36kr": {
            "name": "36氪热榜",
            "columns": ["排名", "标题", "作者", "阅读量", "点赞", "收藏", "评论", "互动分", "engagement_score", "链接"],
            "extract": lambda item, pf="36kr": [
                item.get("rank"), item.get("title"), item.get("author"),
                item.get("read"), item.get("like"), item.get("collect"),
                item.get("comment"),
                *normalize_engagement(pf, item),
                item.get("url")
            ]
        },
        "baidu": {
            "name": "百度热搜",
            "columns": ["排名", "标题", "热度", "描述", "互动分", "engagement_score", "链接"],
            "extract": lambda item, pf="baidu": [
                item.get("rank"), item.get("title"), item.get("heat"),
                item.get("desc"),
                *normalize_engagement(pf, item),
                item.get("url")
            ]
        },
    }

    for platform, config in platform_config.items():
        if platform not in data:
            continue
        result = data[platform]
        if result.get("error"):
            print(f"⚠️ {config['name']}有错误: {result['error']}", file=sys.stderr)
            continue

        ws = wb.create_sheet(config["name"])
        ws.append(config["columns"])

        for item in result.get("items", []):
            ws.append(config["extract"](item))

        # 设置列宽
        for i, col in enumerate(config["columns"]):
            ws.column_dimensions[chr(65 + i)].width = max(8, min(50, len(col) * 4 + 4))

    # 保存
    wb.save(output_path)
    total = sum(data.get(p, {}).get("count", 0) for p in platform_config if p in data)
    print(f"✅ 已保存: {output_path}")
    print(f"   {len(wb.sheetnames)}个平台, 共{total}条数据")
    return output_path

if __name__ == "__main__":
    main()
